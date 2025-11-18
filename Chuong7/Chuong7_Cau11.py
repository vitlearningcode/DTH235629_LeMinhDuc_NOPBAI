# employee_manager_excel.py
from openpyxl import Workbook, load_workbook
import os

class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = int(tuoi)

class QuanLyNhanVien:
    def __init__(self, file_path="nhanvien.xlsx"):
        self.file_path = file_path
        self.ds_nv = []

    # Ghi dữ liệu vào file Excel
    def luu_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        # Tiêu đề cột
        ws.append(["STT", "Mã", "Tên", "Tuổi"])

        # Ghi từng nhân viên
        for i, nv in enumerate(self.ds_nv, start=1):
            ws.append([i, nv.ma, nv.ten, nv.tuoi])

        wb.save(self.file_path)
        print(f"✅ Đã lưu dữ liệu vào '{self.file_path}'")

    # Đọc danh sách nhân viên từ file Excel
    def doc_file(self):
        if not os.path.exists(self.file_path):
            print("⚠️ File chưa tồn tại!")
            return

        wb = load_workbook(self.file_path)
        ws = wb.active

        self.ds_nv.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, ma, ten, tuoi = row
            self.ds_nv.append(NhanVien(ma, ten, tuoi))
        print("✅ Đã đọc dữ liệu từ file Excel!")

    # Thêm nhân viên
    def them_nv(self, nv):
        self.ds_nv.append(nv)

    # Sắp xếp theo tuổi tăng dần
    def sap_xep_tuoi(self):
        self.ds_nv.sort(key=lambda nv: nv.tuoi)
        print("✅ Đã sắp xếp nhân viên theo tuổi tăng dần!")

    # Hiển thị danh sách nhân viên
    def hien_thi(self):
        print(f"\n{'Mã':<10}{'Tên':<15}{'Tuổi':<5}")
        print("-" * 30)
        for nv in self.ds_nv:
            print(f"{nv.ma:<10}{nv.ten:<15}{nv.tuoi:<5}")

# --- Demo ---
if __name__ == "__main__":
    ql = QuanLyNhanVien()

    # Thêm dữ liệu mẫu
    ql.them_nv(NhanVien("NV1", "An", 18))
    ql.them_nv(NhanVien("NV2", "Lành", 22))
    ql.them_nv(NhanVien("NV3", "Giải", 20))
    ql.them_nv(NhanVien("NV4", "Thoát", 19))
    ql.them_nv(NhanVien("NV5", "Hạnh", 25))
    ql.them_nv(NhanVien("NV6", "Phúc", 24))

    # Lưu file
    ql.luu_file()

    # Đọc lại file
    ql.doc_file()

    # Sắp xếp theo tuổi
    ql.sap_xep_tuoi()
    ql.hien_thi()